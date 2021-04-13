#!/usr/bin/env python
# -*- coding: UTF-8 -*-
import os
from openpyxl import load_workbook



# def open_excel(file: str):
# 	"""
# 	:param file: 文件路径s
# 	:return:An instance of the :class:`~xlrd.book.Book` class.
# 	"""
# 	try:
# 		path = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) + file
# 		data = xlrd.open_workbook(path)
# 		return data
# 	except Exception as e:
# 		raise e
#
#
# def excel_table_byname(file: str, by_name: str, colnameindex=0) -> list:
# 	"""
# 	:param file: 文件名
# 	:param by_name: 表sheet页的名称
# 	:param colnameindex: 表头列名所在行的索引
# 	:return data list
# 	"""
# 	data = open_excel(file)  # 打开文件
# 	table = data.sheet_by_name(by_name)  # 通过名称获取sheet页面数据
# 	nrows = table.nrows  # 列数
# 	colnames = table.row_values(colnameindex)  # 第colnameindex行数据
# 	data_list = []  # 准备一个空列表放数据
# 	for rownum in range(1, nrows):  # for循环  由于0是标题，所以循环从1开始
# 		row = table.row_values(rownum)  # row等于rownum行的值
# 		if row:
# 			app = {}  # 定义一个空数组
# 			for i in range(len(colnames)):  # 循环
# 				app[colnames[i]] = row[i]  # 当i有长度时，app数组里面第i行等于row的第i行的值
# 			data_list.append(app)  # 把app的值添加到列表中
# 	return data_list


def open_excel(file: str):
	"""
	:param file: 文件路径
	:return:An instance of the :class:`~xlrd.book.Book` class.
	"""
	try:
		path = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) + file
		data = load_workbook(path)
		return data
	except Exception as e:
		raise e


def excel_table_byname(file: str, by_name: str, colnameindex=1) -> list:
	"""
	:param file: 文件名
	:param by_name: 表sheet页的名称
	:param colnameindex: 表头列名所在行的索引
	:return data list
	"""
	data = open_excel(file)  # 打开文件
	table = data[by_name]  # 通过名称获取sheet页面数据
	n_rows = table.max_row  # 最大行数
	n_cols = table.max_column # 最大列数
	data_list = []  # 准备一个空列表放数据
	for k in range(1, n_rows + 1): # 从第一行开始遍历
		app = {} # 定义一个空字典
		for l in range(1, n_cols + 1): # 从第一列开始遍历
			key = table.cell(row=1, column = l).value # 字典key的名称 excel的 第一行
			value = table.cell(row=k + 1, column=l).value # 字典value的名称 从第二行开始遍历
			if key  is not None: #判断值不为空时加入字典
				if value is None:
					app[key] = ""
				else:
					app[key] = value
		a_set = list(app.values())
		is_none_col = 0
		for s in a_set: # 遍历转成list的字典, 如果字典值为空 则 is_none_col加一
			if s == "":
				is_none_col += 1
		if is_none_col == len(a_set): # 判断is_none_col是否等于app的长度 如果是则代表所有字段为空,转为空字典
			app = {}
		if app != {}: # 判断字典不为空时加入列表
			data_list.append(app)
	return data_list